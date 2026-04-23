import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import base64
import json
from openai import OpenAI

# 페이지 설정
st.set_page_config(page_title="Alibaba Listing Generator", layout="wide")

st.title("🚀 알리바바 상품 등록 자동화 툴 (Streamlit)")
st.markdown("""
이 툴은 상품 스크린샷을 분석하여 알리바바 검색에 최적화된 영어 제목과 키워드를 생성하고, 
즉시 업로드 가능한 엑셀 파일을 만들어줍니다.
""")

# 사이드바 설정 (API 키 입력 등)
with st.sidebar:
    st.header("⚙️ 설정")
    # Streamlit Secrets에서 먼저 찾고, 없으면 입력 받음
    default_api_key = st.secrets.get("OPENAI_API_KEY", "")
    api_key = st.text_input("OpenAI API Key를 입력하세요", value=default_api_key, type="password")
    num_versions = st.slider("제품당 생성할 변형 버전 수", 1, 5, 3)
    target_currency = st.selectbox("통화 단위", ["USD", "KRW", "EUR"], index=0)

# 파일 업로드 부분
st.header("1. 상품 스크린샷 업로드")
uploaded_images = st.file_uploader("상품 상세페이지 스크린샷들을 선택하세요", type=["png", "jpg", "jpeg"], accept_multiple_files=True)

if uploaded_images:
    cols = st.columns(len(uploaded_images))
    for idx, img in enumerate(uploaded_images):
        cols[idx].image(img, use_container_width=True)

# 메인 로직 실행
if st.button("✨ 분석 및 엑셀 생성 시작"):
    if not api_key:
        st.error("OpenAI API Key를 입력해주세요.")
    elif not uploaded_images:
        st.error("이미지를 최소 한 장 이상 업로드해주세요.")
    else:
        with st.spinner("AI가 이미지를 분석하고 최적화된 키워드를 생성 중입니다..."):
            try:
                # 1. AI 분석 및 키워드 추출 (GPT-4o Vision 활용 예시)
                client = OpenAI(api_key=api_key)
                
                # 이미지들을 base64로 변환
                img_payload = []
                for img in uploaded_images:
                    base64_image = base64.b64encode(img.getvalue()).decode('utf-8')
                    img_payload.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}
                    })

                # 프롬프트 구성
                prompt = f"""
                Analyze the attached product screenshots and extract the following information:
                - Product Brand
                - Model Name/Number
                - Key Features (at least 5)
                - Technical Specifications
                - Original Price in KRW (if visible)

                Then, generate {num_versions} different optimized Alibaba Product Titles (max 128 chars each).
                Titles should include high-volume keywords and be attractive to B2B buyers.
                Also, provide a short professional English description (max 1000 chars).
                
                Return the result ONLY as a JSON object with this structure:
                {{
                    "brand": "", "model": "", "features": [], "spec": "", "price_krw": 0,
                    "titles": ["title1", "title2", ...],
                    "description": ""
                }}
                """

                response = client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "user",
                            "content": [{"type": "text", "text": prompt}] + img_payload
                        }
                    ],
                    response_format={ "type": "json_object" }
                )
                
                result = json.loads(response.choices[0].message.content)
                st.success("✅ 분석 완료!")
                
                # 분석 결과 표시
                with st.expander("🔍 분석된 상품 정보 확인"):
                    st.json(result)

                # 2. 엑셀 파일 생성
                template_path = "template.xlsx" # 미리 준비된 템플릿 파일
                wb = load_workbook(template_path)
                ws = wb.active
                
                start_row = 5
                for i, title in enumerate(result['titles']):
                    row = start_row + i
                    ws.cell(row=row, column=1, value=title) # Title
                    ws.cell(row=row, column=8, value=result['description']) # Description
                    ws.cell(row=row, column=10, value=result['brand']) # Brand
                    ws.cell(row=row, column=15, value=result['model']) # Model
                    ws.cell(row=row, column=25, value=target_currency) # Currency
                    ws.cell(row=row, column=26, value='Range pricing')
                    
                    # 가격 변환 (대략적 환율 적용 또는 직접 입력 가능하게 확장 가능)
                    price_usd = round(result.get('price_krw', 0) / 1300, 2)
                    ws.cell(row=row, column=45, value=price_usd)
                    ws.cell(row=row, column=46, value=price_usd * 1.2)

                # 메모리에 엑셀 저장
                excel_out = io.BytesIO()
                wb.save(excel_out)
                excel_out.seek(0)
                
                st.download_button(
                    label="📥 최적화된 알리바바 엑셀 다운로드",
                    data=excel_out,
                    file_name=f"Alibaba_Optimized_{result['model']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"오류가 발생했습니다: {e}")

st.markdown("---")
st.caption("Accio Work Assistant - Alibaba Automation Project")
